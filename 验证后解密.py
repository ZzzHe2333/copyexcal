import tkinter as tk
from tkinter import filedialog
import openpyxl
import datetime
from time import sleep
import hashlib

passwords = {"12345678"} # 初始化密码库，初始密码为"12345678"


# 获取本地时间并格式化为"年-月-日-时-分"
def get_formatted_local_time():
    now = datetime.datetime.now()
    formatted_time = now.strftime("%Y%m%d%H%M")
    return formatted_time

# 计算字体大小，要求是屏幕宽度的1/3
def calculate_font_size():
    screen_width = root.winfo_screenwidth()  # 获取屏幕宽度
    dpi = root.winfo_fpixels("1i")  # 获取DPI
    font_size = int(screen_width / 3 / dpi * 10)  # 计算字体大小
    return font_size

def get_password_hash(user_input):
    try:
        number = int(user_input)
        number = str(number)
    # 计算 SHA-256 哈希值
        sha256_hash = hashlib.sha256(number.encode('utf-8')).hexdigest()
        #print(f"输入整数的SHA-256哈希值: {sha256_hash}")
        return sha256_hash
    except ValueError:
        print("输入无效的整数")

def get_password_md5(user_input):
    try:
        number = int(user_input)
        number = str(number)
    #计算md5
        md5 = hashlib.md5(number.encode('utf-8')).hexdigest()
        #print(f"输入整数的MD5哈希值: {md5_hash}")
        return md5
    except ValueError:
        print("输入无效的整数")

# 弹出提示窗口并自动关闭,含参数
def show_completion_message_diy(message:str,time=2000,):
    completion_window = tk.Toplevel(root)
    completion_window.geometry("500x100")  # 设置窗口大小
    font_size = calculate_font_size()
    completion_label = tk.Label(completion_window, text=message, font=("verdana", font_size,"bold"), justify="center")
    completion_label.pack(expand=True, fill='both')
    # 设置2秒后自动关闭,默认值2000，就是2s。
    root.after(time, completion_window.destroy)


# 检查密码是否在密码库中
def check_password(password):
    print("----------\n验证程序\n----------")
    return password in passwords

def show_current_time():
    current_time = get_formatted_local_time()
    
    time_window = tk.Toplevel(root)
    time_label = tk.Label(time_window, text=f"当前时间: {current_time}")
    time_label.pack()
    root.after(2000, time_window.destroy)
    print(f"当前时间: {current_time}")
    
# 打开并复制xlsx文件的回调函数
def open_and_copy_xlsx():
    #宣告全局变量
    global passwords
    # 初始化密码库，初始密码为""
    passwords = {"ab210dc0bc203a457e33b77d048b9b2044d11f459540195ecdf75dd86388c0ba"} 
    #新增密码
    passwords.add("093c3c520aee12782d6dcce9a32050d39585da1639c0704e26c08e1831e1bb5f")
    passwords.add("345ebf3b583cacca53b7e244a5ffae3ec96578558984d2224a2a9461f45c6337") 
    passwords.add("b4e93923d13dc98f822d27d1fa2b48639a628a9c61bf68150b5a0956269eaba8")


    user_input_password = password_entry.get()  # 获取用户输入的密码
    use_clone_file = clone_var.get()  # 获取复选框的状态
    use_close = close_var.get()

#添加hash值进入密码库，
    xunhuan_x = -6 #循环标志符
    xunhuan_firsttime = get_formatted_local_time() #获取当地时间 202309011800 年月日时分
    xunhuan_chushi = int(xunhuan_firsttime) - 5 #前5分钟hash值有效
    while xunhuan_x < 0:
        passwords.add(get_password_hash(xunhuan_chushi))
        print (get_password_hash(xunhuan_chushi))
        xunhuan_chushi += 1
        xunhuan_x += 1     

    #会生成7个值

#添加md5值进入密码库，
    xunhuan_y = -4 #循环标志符
    xunhuan_firsttime = get_formatted_local_time() #获取当地时间 202309011800 年月日时分
    xunhuan_chushi = int(xunhuan_firsttime) - 3 #前3分钟的md5值有效
    while xunhuan_y < 0:
        passwords.add(get_password_md5(xunhuan_chushi))
        print (get_password_md5(xunhuan_chushi))
        xunhuan_chushi += 1
        xunhuan_y += 1
        pass
    # print (passwords) #debug的时候用
    # 检查密码是否在密码库中
    if check_password(user_input_password):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            # 打开选定的.xlsx文件
            wb = openpyxl.load_workbook(file_path)
            all_data = []  # 用于存储所有工作表的数据

            # 遍历每一页工作表
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                # 读取数据
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(row)

                # 将数据添加到所有工作表数据中
                all_data.append((sheet_name, data))

            # 创建新的.xlsx文件
            if use_clone_file:
                # 如果复选框被选中，生成xxxx-clone.xlsx文件
                new_file_path = file_path.replace('.xlsx', '-clone.xlsx')
            else:
                # 否则，使用用户指定的文件名
                new_file_path = filedialog.asksaveasfilename(filetypes=[("Excel Files", "*.xlsx")], defaultextension=".xlsx")

            if new_file_path:
                new_wb = openpyxl.Workbook()

                # 遍历所有工作表的数据并写入新文件
                for sheet_name, data in all_data:
                    new_sheet = new_wb.create_sheet(title=sheet_name)
                    for row_data in data:
                        new_sheet.append(row_data)

                # 删除新文件中默认的Sheet
                default_sheet = new_wb['Sheet']
                new_wb.remove(default_sheet)

                # 保存新文件
                new_wb.save(new_file_path)

                # 关闭新文件
                new_wb.close()

            # 关闭原文件
            wb.close()

            # 复制完成后，弹出提示窗口并自动关闭
            show_completion_message_diy("紧急解密完成",time=2000)
            if use_close is True:
                root.after(2000, root.destroy)

    else:
        if user_input_password is "":
            show_completion_message_diy("未输入授权码",time=1000)
        else:
            try:
                popopopo = int(user_input_password)
            except ValueError :
                show_completion_message_diy("授权码错误",time=1500)

# 调用tellboss函数的壳子
def tellboss():
    pass

#关闭窗口
def close_UI():
    root.destroy()
#-------------------------------------------------------------

# 创建GUI窗口
root = tk.Tk()
root.title("[紧急用]excal解密器")
root.geometry("300x300")  # 设置窗口大小为600x250像素

# 创建一个Frame来容纳控件，并使其在窗口中居中
frame = tk.Frame(root,padx=5,pady=5)
frame.place(relx=0.5, rely=0.5, anchor="center")

# 创建一个新的输入框用于用户输入密码
password_label = tk.Label(frame, text="请输入授权码")
password_label.pack()
password_entry = tk.Entry(frame, show="")  # 使用*显示密码字符
password_entry.pack(pady=2)

# 创建一个复选框，控制生成方式
clone_var = tk.BooleanVar(value=True)
clone_checkbox = tk.Checkbutton(frame, text="生成-clone.xlsx文件", variable=clone_var)
clone_checkbox.pack(pady=5)

# 创建一个复选框，控制生成方式
close_var = tk.BooleanVar()
close_checkbox = tk.Checkbutton(frame, text="执行完顺带关程序", variable=close_var)
close_checkbox.pack(pady=3)

# 创建按钮，并将其放在Frame中
open_button = tk.Button(frame, text="紧急解密xlsx文件", command=open_and_copy_xlsx)
open_button.pack(pady=5,ipadx=20,ipady=5,fill="both")

# 创建一个按钮，点击显示当前时间
show_time_button = tk.Button(frame, text="显示当前时间", command=show_current_time)
show_time_button.pack(pady=5,ipadx=20,ipady=5,fill="both")

# 创建一个按钮，关闭程序
close_button = tk.Button(frame, text="关闭", command=close_UI)
close_button.pack(pady=15,ipadx=20,ipady=5,fill="both")


# 运行GUI程序
root.mainloop()
